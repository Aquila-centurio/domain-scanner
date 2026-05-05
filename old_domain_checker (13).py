import csv
import io
import time
import sys
import threading
from datetime import datetime, timezone, timedelta
from pathlib import Path

from flask import Flask, request, jsonify, send_file
import requests as http_requests
import openpyxl

app = Flask(__name__)

RAPIDAPI_KEY = "db6d4987ecmshd4a7aa2c1fc0ca7p1b4c58jsn7cc388f1d736"
VT_API_KEY = "0be1f79cdcaf7a46a9c9c86e4007ff64aad993d5b5a89897d795bec0934f32be"

scan_state = {
    "running": False, "logs": [], "results_great": [], "results_good": [], "flagged": [],
    "stats": {}, "progress": 0, "total": 0, "step": "",
}


def reset_state():
    scan_state.update({"running": False, "logs": [], "results_great": [], "results_good": [],
                        "flagged": [], "stats": {}, "progress": 0, "total": 0, "step": ""})


def log(msg, level="info"):
    scan_state["logs"].append({"msg": msg, "level": level, "time": datetime.now().strftime("%H:%M:%S")})


def parse_rows(file_bytes, filename):
    ext = Path(filename).suffix.lower() if filename else ""
    if ext in (".csv", ".tsv"):
        text = file_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        next(reader, None)
        for row in reader:
            if len(row) >= 15:
                yield row
            elif len(row) == 1 and "," in row[0]:
                parts = row[0].split(",")
                if len(parts) >= 15:
                    yield parts
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[0]
            if not val or not isinstance(val, str):
                continue
            parts = val.split(",")
            if len(parts) >= 15:
                yield parts
            elif len(row) >= 15:
                yield [str(c) if c is not None else "" for c in row]
        wb.close()


def load_and_filter(file_bytes, min_hours, max_hours, max_price, filename=""):
    now = datetime.now(timezone.utc)
    window_start = now + timedelta(hours=min_hours)
    window_end = now + timedelta(hours=max_hours)
    domains = []
    total = passed_com = passed_auction = passed_price = passed_reg = 0

    for parts in parse_rows(file_bytes, filename):
        total += 1
        domain = parts[1].strip().lower()
        if not domain.endswith(".com"):
            continue
        passed_com += 1

        reg_date = parts[14].strip() if len(parts) > 14 else ""
        reg_year = 0
        if reg_date:
            try:
                reg_year = int(reg_date[:4])
            except (ValueError, IndexError):
                pass
        if reg_year < 2000 or reg_year > 2016:
            continue
        passed_reg += 1

        end_str = parts[3].strip()
        if not end_str:
            continue
        try:
            end_dt = datetime.fromisoformat(end_str.replace("Z", "+00:00"))
        except ValueError:
            continue
        if end_dt < window_start or end_dt > window_end:
            continue
        passed_auction += 1

        def sf(idx):
            try:
                return float(parts[idx])
            except (IndexError, ValueError, TypeError):
                return 0.0

        price = sf(4)
        if max_price > 0 and price > max_price:
            continue
        passed_price += 1

        domains.append({
            "domain": domain, "reg_date": reg_date[:10], "reg_year": reg_year,
            "end_date": end_str[:19].replace("T", " "),
            "hours_left": round((end_dt - now).total_seconds() / 3600, 1),
            "price": price, "ahrefs_dr": sf(8), "majestic_tf": sf(23),
            "backlinks": sf(20), "bid_count": int(sf(7)), "url": parts[0].strip(),
        })

    domains.sort(key=lambda x: x["hours_left"])
    scan_state["stats"] = {"total_rows": total, "dot_com": passed_com,
                            "reg_2000_2016": passed_reg, "auction_window": passed_auction, "price_ok": passed_price}
    return domains


def check_scamdoc(domain, retries=3):
    """API returns final_score as risk (0-1). Trust = (1 - risk) * 100.
    Example: final_score 0.14 means trust 86% (matches scamdoc.com)."""
    url = "https://scampredictor.p.rapidapi.com/domain/" + domain
    headers = {
        "x-rapidapi-key": RAPIDAPI_KEY,
        "x-rapidapi-host": "scampredictor.p.rapidapi.com",
        "Content-Type": "application/json",
    }
    for attempt in range(retries):
        try:
            timeout = 45 + (attempt * 15)
            resp = http_requests.get(url, headers=headers, timeout=timeout)
            if resp.status_code == 200:
                data = resp.json()
                final_score = data.get("final_score")
                if final_score is not None:
                    risk = float(final_score)
                    trust = max(0, min(100, round((1 - risk) * 100)))
                    log("ScamDoc [" + domain + "]: risk=" + str(risk) + " trust=" + str(trust) + "%", "dim")
                    return trust
                log("ScamDoc [" + domain + "]: no final_score in " + str(data)[:200], "warning")
                return None
            elif resp.status_code == 429:
                log("ScamDoc rate limited. Waiting 15s...", "warning")
                time.sleep(15)
                continue
            else:
                log("ScamDoc [" + domain + "]: HTTP " + str(resp.status_code), "warning")
                return None
        except Exception as e:
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                log("ScamDoc timeout for " + domain + ", retry " + str(attempt + 2) + "/" + str(retries) + " in " + str(wait) + "s...", "warning")
                time.sleep(wait)
            else:
                log("ScamDoc failed after " + str(retries) + " tries: " + domain, "error")
    return None


def check_virustotal(domain):
    headers = {"x-apikey": VT_API_KEY}
    try:
        resp = http_requests.get("https://www.virustotal.com/api/v3/domains/" + domain,
                                  headers=headers, timeout=30)
        if resp.status_code == 200:
            stats = resp.json().get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            return {"malicious": stats.get("malicious", 0), "suspicious": stats.get("suspicious", 0),
                    "clean": stats.get("malicious", 0) == 0 and stats.get("suspicious", 0) == 0}
        elif resp.status_code == 429:
            time.sleep(60)
            return check_virustotal(domain)
        elif resp.status_code == 404:
            return {"malicious": 0, "suspicious": 0, "clean": True}
    except Exception:
        pass
    return {"malicious": 0, "suspicious": 0, "clean": True}


def run_scan(file_bytes, min_hours, max_hours, limit, min_sd_score, max_sd_score, max_price, filename=""):
    reset_state()
    scan_state["running"] = True
    scan_state["step"] = "filtering"

    log("Loading and filtering: " + filename)
    domains = load_and_filter(file_bytes, min_hours, max_hours, max_price, filename)

    s = scan_state["stats"]
    log("Total rows: " + str(s["total_rows"]))
    log(".com domains: " + str(s["dot_com"]))
    log("Registered 2000-2016: " + str(s["reg_2000_2016"]))
    log("Auction " + str(min_hours) + "-" + str(max_hours) + "h: " + str(s["auction_window"]))
    log("Price filter passed: " + str(s["price_ok"]), "success")

    if not domains:
        log("No domains match filters. Try widening the auction window.", "error")
        scan_state["running"] = False
        scan_state["step"] = "done"
        return

    if limit > 0 and limit < len(domains):
        domains = domains[:limit]
        log("Limited to first " + str(limit) + " domains.")

    scan_state["total"] = len(domains)

    # Step 2: ScamDoc
    scan_state["step"] = "scamdoc"
    log("--- ScamDoc scan (" + str(len(domains)) + " domains) ---")

    passed_sd = []
    for i, entry in enumerate(domains):
        if not scan_state["running"]:
            break
        scan_state["progress"] = i + 1
        domain = entry["domain"]
        score = check_scamdoc(domain)

        if score is not None:
            entry["scamdoc_score"] = score
            if score >= min_sd_score:
                tier = "GREAT" if score >= max_sd_score else "GOOD"
                passed_sd.append(entry)
                log(tier + " " + domain + ": " + str(score) + "% | " + str(entry["hours_left"]) + "h | $" + str(entry["price"]), "success")
            else:
                log("SKIP " + domain + ": " + str(score) + "% (below " + str(min_sd_score) + ")", "dim")
        else:
            log("? " + domain + ": no score returned", "warning")
        time.sleep(1.5)

    if not passed_sd:
        log("No domains scored " + str(min_sd_score) + "%+.", "warning")
        scan_state["running"] = False
        scan_state["step"] = "done"
        return

    log("--- " + str(len(passed_sd)) + " domains passed ScamDoc ---", "success")

    # Step 3: VirusTotal
    scan_state["step"] = "virustotal"
    scan_state["progress"] = 0
    scan_state["total"] = len(passed_sd)
    log("--- VirusTotal scan (" + str(len(passed_sd)) + " domains) ---")

    for i, entry in enumerate(passed_sd):
        if not scan_state["running"]:
            break
        scan_state["progress"] = i + 1
        domain = entry["domain"]
        vt = check_virustotal(domain)
        entry["vt_malicious"] = vt["malicious"]
        entry["vt_suspicious"] = vt["suspicious"]

        if vt["clean"]:
            if entry["scamdoc_score"] >= max_sd_score:
                scan_state["results_great"].append(entry)
                log("CLEAN+GREAT " + domain + " (SD: " + str(entry["scamdoc_score"]) + "%)", "success")
            else:
                scan_state["results_good"].append(entry)
                log("CLEAN+GOOD " + domain + " (SD: " + str(entry["scamdoc_score"]) + "%)", "success")
        else:
            scan_state["flagged"].append(entry)
            log("FLAGGED " + domain + " (" + str(vt["malicious"]) + " malicious)", "error")
        time.sleep(16)

    great = len(scan_state["results_great"])
    good = len(scan_state["results_good"])
    flagged = len(scan_state["flagged"])
    log("--- DONE: " + str(great) + " great, " + str(good) + " good, " + str(flagged) + " flagged ---", "success" if (great + good) > 0 else "warning")
    scan_state["step"] = "done"
    scan_state["running"] = False


# ===== Routes =====

@app.route("/")
def index():
    return get_html()


@app.route("/scan", methods=["POST"])
def start_scan():
    if scan_state["running"]:
        return jsonify({"error": "Scan already running"}), 400
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    min_h = float(request.form.get("min_hours", 6))
    max_h = float(request.form.get("max_hours", 24))
    limit = int(request.form.get("limit", 50))
    min_sd = int(request.form.get("min_sd_score", 80))
    max_sd = int(request.form.get("max_sd_score", 100))
    max_price = float(request.form.get("max_price", 0))
    file_bytes = f.read()
    filename = f.filename or ""
    thread = threading.Thread(target=run_scan, args=(file_bytes, min_h, max_h, limit, min_sd, max_sd, max_price, filename))
    thread.daemon = True
    thread.start()
    return jsonify({"status": "started"})


@app.route("/status")
def status():
    return jsonify({
        "running": scan_state["running"], "step": scan_state["step"],
        "progress": scan_state["progress"], "total": scan_state["total"],
        "logs": scan_state["logs"][-50:], "results_great": scan_state["results_great"],
        "results_good": scan_state["results_good"],
        "flagged": scan_state["flagged"], "stats": scan_state["stats"],
    })


@app.route("/stop", methods=["POST"])
def stop_scan():
    scan_state["running"] = False
    return jsonify({"status": "stopping"})


@app.route("/download")
def download_csv():
    if not scan_state["results_great"] and not scan_state["results_good"]:
        return "No results", 404

    fields = ["domain", "scamdoc_score", "hours_left", "end_date", "price",
              "bid_count", "reg_date", "ahrefs_dr", "backlinks",
              "vt_malicious", "vt_suspicious", "url"]

    wb = openpyxl.Workbook()

    # Sheet 1: Great domains (above max threshold)
    ws1 = wb.active
    ws1.title = "Great Domains"
    ws1.append(fields)
    for d in sorted(scan_state["results_great"], key=lambda x: x.get("scamdoc_score", 0), reverse=True):
        ws1.append([d.get(f, "") for f in fields])

    # Sheet 2: Good domains (between min and max threshold)
    ws2 = wb.create_sheet("Good Domains")
    ws2.append(fields)
    for d in sorted(scan_state["results_good"], key=lambda x: x.get("scamdoc_score", 0), reverse=True):
        ws2.append([d.get(f, "") for f in fields])

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="domain_results.xlsx")


def get_html():
    return (
        '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">'
        '<title>Domain Scanner</title>'
        '<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;600;700&family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">'
        '<style>'
        '*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}'
        ':root{'
        '--bg:#0a0e17;--bg2:#111827;--bg3:#1a2234;--border:#1e2d45;'
        '--text:#c9d1d9;--text-dim:#6b7b8d;--accent:#00e59b;--accent2:#00c4ff;'
        '--danger:#ff4757;--warning:#ffa502;--success:#00e59b;'
        '}'
        'body{font-family:Outfit,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}'
        'body::before{content:"";position:fixed;inset:0;'
        'background:radial-gradient(ellipse at 20% 50%,rgba(0,229,155,.03)0%,transparent 50%),'
        'radial-gradient(ellipse at 80% 20%,rgba(0,196,255,.03)0%,transparent 50%);pointer-events:none}'
        '.container{max-width:1100px;margin:0 auto;padding:40px 24px;position:relative;z-index:1}'
        'header{text-align:center;margin-bottom:48px}'
        'header h1{font-size:2.4rem;font-weight:800;letter-spacing:-1px;'
        'background:linear-gradient(135deg,var(--accent),var(--accent2));'
        '-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:8px}'
        'header p{color:var(--text-dim);font-size:.95rem;font-family:JetBrains Mono,monospace;font-weight:300}'
        '.card{background:linear-gradient(135deg,#111827,#0d1321);border:1px solid var(--border);'
        'border-radius:16px;padding:32px;margin-bottom:24px}'
        '.card h2{font-size:1.1rem;font-weight:600;margin-bottom:24px;color:var(--accent);display:flex;align-items:center;gap:10px}'
        '.card h2 .dot{width:8px;height:8px;border-radius:50%;background:var(--accent);display:inline-block;box-shadow:0 0 12px var(--accent)}'
        '.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}'
        '@media(max-width:640px){.form-grid{grid-template-columns:1fr}}'
        '.field{display:flex;flex-direction:column;gap:8px}.field.full{grid-column:1/-1}'
        'label{font-size:.82rem;font-weight:500;color:var(--text-dim);text-transform:uppercase;letter-spacing:1px;font-family:JetBrains Mono,monospace}'
        'input[type="number"],input[type="file"]{background:var(--bg);border:1px solid var(--border);'
        'border-radius:10px;padding:12px 16px;color:var(--text);font-family:JetBrains Mono,monospace;font-size:.95rem;outline:none;'
        'transition:border-color .2s,box-shadow .2s}'
        'input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,229,155,.1)}'
        'input[type="file"]{cursor:pointer;padding:14px 16px}'
        'input[type="file"]::file-selector-button{background:linear-gradient(135deg,var(--accent),var(--accent2));'
        'color:#000;border:none;padding:8px 18px;border-radius:8px;font-family:Outfit,sans-serif;font-weight:600;font-size:.85rem;cursor:pointer;margin-right:12px}'
        '.actions{display:flex;gap:12px;margin-top:28px;grid-column:1/-1}'
        'button{font-family:Outfit,sans-serif;font-weight:600;font-size:.95rem;border:none;border-radius:10px;padding:14px 32px;cursor:pointer;transition:all .2s}'
        '.btn-go{background:linear-gradient(135deg,var(--accent),var(--accent2));color:#000;flex:1;box-shadow:0 4px 24px rgba(0,229,155,.2)}'
        '.btn-go:hover{transform:translateY(-1px);box-shadow:0 6px 32px rgba(0,229,155,.3)}'
        '.btn-go:disabled{opacity:.4;cursor:not-allowed;transform:none}'
        '.btn-stop{background:rgba(255,71,87,.15);color:var(--danger);border:1px solid rgba(255,71,87,.3)}'
        '.btn-stop:hover{background:rgba(255,71,87,.25)}'
        '.btn-csv{background:rgba(0,229,155,.1);color:var(--accent);border:1px solid rgba(0,229,155,.2)}'
        '.btn-csv:hover{background:rgba(0,229,155,.2)}'
        '.progress-bar{height:4px;background:var(--bg);border-radius:4px;margin-bottom:16px;overflow:hidden}'
        '.progress-fill{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:4px;transition:width .4s;box-shadow:0 0 16px var(--accent)}'
        '.step-label{font-family:JetBrains Mono,monospace;font-size:.8rem;color:var(--accent);margin-bottom:12px;display:flex;justify-content:space-between}'
        '.console{background:#050810;border:1px solid var(--border);border-radius:12px;padding:16px;'
        'max-height:320px;overflow-y:auto;font-family:JetBrains Mono,monospace;font-size:.78rem;line-height:1.8;scroll-behavior:smooth}'
        '.console::-webkit-scrollbar{width:6px}.console::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px}'
        '.log-line{display:flex;gap:10px}.log-time{color:var(--text-dim);flex-shrink:0}.log-msg{word-break:break-all}'
        '.log-line.success .log-msg{color:var(--success)}'
        '.log-line.error .log-msg{color:var(--danger)}'
        '.log-line.warning .log-msg{color:var(--warning)}'
        '.log-line.dim .log-msg{color:var(--text-dim)}'
        '.results-table{width:100%;border-collapse:collapse;font-size:.85rem}'
        '.results-table th{text-align:left;padding:12px 14px;font-family:JetBrains Mono,monospace;'
        'font-size:.72rem;font-weight:500;color:var(--text-dim);text-transform:uppercase;letter-spacing:1px;'
        'border-bottom:1px solid var(--border);position:sticky;top:0;background:var(--bg2)}'
        '.results-table td{padding:11px 14px;border-bottom:1px solid rgba(30,45,69,.5);transition:background .15s}'
        '.results-table tr:hover td{background:rgba(0,229,155,.03)}'
        '.domain-name{font-family:JetBrains Mono,monospace;font-weight:500;color:var(--accent2)}'
        '.score-badge{display:inline-block;padding:3px 10px;border-radius:6px;font-family:JetBrains Mono,monospace;font-weight:600;font-size:.82rem}'
        '.score-high{background:rgba(0,229,155,.15);color:var(--accent)}'
        '.score-med{background:rgba(255,165,2,.15);color:var(--warning)}'
        '.vt-clean{color:var(--success)}.vt-flagged{color:var(--danger);font-weight:600}'
        '.table-wrap{max-height:500px;overflow-y:auto;border-radius:12px;border:1px solid var(--border)}'
        '.table-wrap::-webkit-scrollbar{width:6px}.table-wrap::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px}'
        '.stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px}'
        '@media(max-width:640px){.stat-grid{grid-template-columns:repeat(2,1fr)}}'
        '.stat-box{background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:14px 16px;text-align:center}'
        '.stat-num{font-family:JetBrains Mono,monospace;font-size:1.6rem;font-weight:700;color:var(--accent)}'
        '.stat-label{font-size:.7rem;color:var(--text-dim);text-transform:uppercase;letter-spacing:1px;margin-top:4px}'
        '.hidden{display:none}'
        '@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}'
        '.scanning .dot{animation:pulse 1.2s infinite}'
        '.bid-link{color:var(--accent2);text-decoration:none;font-size:.78rem;opacity:.7;transition:opacity .2s}'
        '.bid-link:hover{opacity:1;text-decoration:underline}'
        '</style></head><body><div class="container">'
        '<header><h1>Domain Scanner</h1>'
        '<p>upload .xlsx or .csv &#8594; auction filter &#8594; scamdoc &#8594; virustotal</p></header>'
        '<div class="card"><h2><span class="dot"></span> Scan Settings</h2>'
        '<div class="form-grid">'
        '<div class="field full"><label>Namecheap File (.xlsx or .csv)</label>'
        '<input type="file" id="fileInput" accept=".xlsx,.xls,.csv,.tsv"></div>'
        '<div class="field"><label>Auction Ends (min hours)</label>'
        '<input type="number" id="minHours" value="3" min="0" step="1"></div>'
        '<div class="field"><label>Auction Ends (max hours)</label>'
        '<input type="number" id="maxHours" value="24" min="1" step="1"></div>'
        '<div class="field"><label>Min ScamDoc Score (skip below)</label>'
        '<input type="number" id="minSD" value="70" min="0" max="100"></div>'
        '<div class="field"><label>Great Threshold (above = great)</label>'
        '<input type="number" id="maxSD" value="80" min="0" max="100"></div>'
        '<div class="field"><label>Max Price $ (0=any)</label>'
        '<input type="number" id="maxPrice" value="20" min="0" step="1"></div>'
        '<div class="field"><label>Max Domains to Check (0=all)</label>'
        '<input type="number" id="limit" value="50" min="0"></div>'
        '<div class="actions">'
        '<button class="btn-go" id="btnScan" onclick="startScan()">Start Scan</button>'
        '<button class="btn-stop hidden" id="btnStop" onclick="stopScan()">Stop</button>'
        '</div></div></div>'
        '<div class="card hidden" id="progressCard">'
        '<h2 class="scanning"><span class="dot"></span> <span id="stepLabel">Scanning...</span></h2>'
        '<div class="step-label"><span id="progressText">0 / 0</span><span id="progressPct">0%</span></div>'
        '<div class="progress-bar"><div class="progress-fill" id="progressFill" style="width:0%"></div></div>'
        '<div class="console" id="logConsole"></div></div>'
        '<div class="card hidden" id="resultsGreatCard">'
        '<h2><span class="dot" style="background:var(--accent)"></span> Great Domains (above threshold)</h2>'
        '<div class="stat-grid" id="statGrid"></div>'
        '<div class="table-wrap"><table class="results-table"><thead><tr>'
        '<th>#</th><th>Domain</th><th>ScamDoc</th><th>Ends In</th>'
        '<th>Price</th><th>Bids</th><th>Registered</th><th>DR</th><th>VT</th><th></th>'
        '</tr></thead><tbody id="greatBody"></tbody></table></div></div>'
        '<div class="card hidden" id="resultsGoodCard">'
        '<h2><span class="dot" style="background:var(--warning)"></span> Good Domains (between min and max)</h2>'
        '<div class="table-wrap"><table class="results-table"><thead><tr>'
        '<th>#</th><th>Domain</th><th>ScamDoc</th><th>Ends In</th>'
        '<th>Price</th><th>Bids</th><th>Registered</th><th>DR</th><th>VT</th><th></th>'
        '</tr></thead><tbody id="goodBody"></tbody></table></div></div>'
        '<div class="card hidden" id="downloadCard">'
        '<div style="display:flex;gap:12px">'
        '<button class="btn-csv" onclick="downloadCSV()">Download Results (.xlsx)</button></div></div>'
        '</div><script>'
        'let pollTimer=null,lastLogCount=0;'
        'function startScan(){'
        'const fi=document.getElementById("fileInput");'
        'if(!fi.files.length){alert("Select a file first.");return}'
        'const fd=new FormData();'
        'fd.append("file",fi.files[0]);'
        'fd.append("min_hours",document.getElementById("minHours").value);'
        'fd.append("max_hours",document.getElementById("maxHours").value);'
        'fd.append("limit",document.getElementById("limit").value);'
        'fd.append("min_sd_score",document.getElementById("minSD").value);'
        'fd.append("max_sd_score",document.getElementById("maxSD").value);'
        'fd.append("max_price",document.getElementById("maxPrice").value);'
        'document.getElementById("btnScan").disabled=true;'
        'document.getElementById("btnStop").classList.remove("hidden");'
        'document.getElementById("progressCard").classList.remove("hidden");'
        'document.getElementById("resultsGreatCard").classList.add("hidden");'
        'document.getElementById("resultsGoodCard").classList.add("hidden");'
        'document.getElementById("downloadCard").classList.add("hidden");'
        'document.getElementById("logConsole").innerHTML="";'
        'document.getElementById("greatBody").innerHTML="";'
        'document.getElementById("goodBody").innerHTML="";'
        'lastLogCount=0;'
        'fetch("/scan",{method:"POST",body:fd})'
        '.then(r=>r.json()).then(d=>{if(d.error){alert(d.error);resetUI();return}'
        'pollTimer=setInterval(pollStatus,1500)}).catch(e=>{alert("Error: "+e);resetUI()})}'
        'function stopScan(){fetch("/stop",{method:"POST"})}'
        'function pollStatus(){'
        'fetch("/status").then(r=>r.json()).then(data=>{'
        'const pct=data.total>0?Math.round((data.progress/data.total)*100):0;'
        'document.getElementById("progressFill").style.width=pct+"%";'
        'document.getElementById("progressPct").textContent=pct+"%";'
        'document.getElementById("progressText").textContent=data.progress+" / "+data.total;'
        'const steps={filtering:"Filtering...",scamdoc:"ScamDoc Scan",virustotal:"VirusTotal Scan",done:"Complete"};'
        'document.getElementById("stepLabel").textContent=steps[data.step]||data.step;'
        'if(data.logs.length>lastLogCount){'
        'const c=document.getElementById("logConsole");'
        'data.logs.slice(lastLogCount).forEach(l=>{'
        'c.innerHTML+=\'<div class="log-line \'+l.level+\'"><span class="log-time">\'+l.time+\'</span><span class="log-msg">\'+escHtml(l.msg)+\'</span></div>\'});'
        'c.scrollTop=c.scrollHeight;lastLogCount=data.logs.length}'
        'if(data.results_great.length>0||data.results_good.length>0||data.flagged.length>0)'
        'renderResults(data.results_great,data.results_good,data.flagged,data.stats);'
        'if(!data.running&&data.step==="done"){clearInterval(pollTimer);resetUI()}})}'
        'function buildRows(list){'
        'const sorted=list.sort((a,b)=>(b.scamdoc_score||0)-(a.scamdoc_score||0));'
        'return sorted.map((d,i)=>{'
        'const sc=d.scamdoc_score||0;const cls=sc>=80?"score-high":"score-med";'
        'const vtCls=(d.vt_malicious||0)===0?"vt-clean":"vt-flagged";'
        'const vtTxt=(d.vt_malicious||0)===0?"Clean":d.vt_malicious+" flags";'
        'const price=d.price>0?"$"+d.price:"-";'
        'const dr=d.ahrefs_dr>0?Math.round(d.ahrefs_dr):"-";'
        'return "<tr><td style=\\"color:var(--text-dim)\\">"+(i+1)+"</td>"+'
        '"<td class=\\"domain-name\\">"+escHtml(d.domain)+"</td>"+'
        '"<td><span class=\\"score-badge "+cls+"\\">"+sc+"%</span></td>"+'
        '"<td>"+d.hours_left+"h</td>"+'
        '"<td style=\\"color:var(--warning)\\">"+price+"</td>"+'
        '"<td>"+(d.bid_count||0)+"</td>"+'
        '"<td style=\\"color:var(--text-dim)\\">"+(d.reg_date||"")+"</td>"+'
        '"<td>"+dr+"</td>"+'
        '"<td class=\\""+vtCls+"\\">"+vtTxt+"</td>"+'
        '"<td><a href=\\""+escHtml(d.url||"#")+"\\" target=\\"_blank\\" class=\\"bid-link\\">Bid &#8594;</a></td></tr>"'
        '}).join("")}'
        'function renderResults(great,good,flagged,stats){'
        'document.getElementById("statGrid").innerHTML='
        'statBox(great.length,"Great")+statBox(good.length,"Good")+'
        'statBox(flagged.length,"VT Flagged")+statBox(stats.reg_2000_2016||0,"Reg 00-16");'
        'if(great.length>0){'
        'document.getElementById("resultsGreatCard").classList.remove("hidden");'
        'document.getElementById("greatBody").innerHTML=buildRows(great)}'
        'if(good.length>0){'
        'document.getElementById("resultsGoodCard").classList.remove("hidden");'
        'document.getElementById("goodBody").innerHTML=buildRows(good)}'
        'if(great.length>0||good.length>0)document.getElementById("downloadCard").classList.remove("hidden")}'
        'function statBox(n,l){return \'<div class="stat-box"><div class="stat-num">\'+n+\'</div><div class="stat-label">\'+l+\'</div></div>\'}'
        'function resetUI(){document.getElementById("btnScan").disabled=false;document.getElementById("btnStop").classList.add("hidden")}'
        'function downloadCSV(){window.location.href="/download"}'
        'function escHtml(s){const d=document.createElement("div");d.textContent=s;return d.innerHTML}'
        '</script></body></html>'
    )


if __name__ == "__main__":
    print("")
    print("  Domain Scanner is running!")
    print("  Open http://localhost:8080 in your browser")
    print("  Press Ctrl+C to stop")
    print("")
    app.run(host="127.0.0.1", port=8080, debug=False)
